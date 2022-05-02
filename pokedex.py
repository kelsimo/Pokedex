from tkinter import *
from tkinter.ttk import *
from openpyxl import workbook, load_workbook
from PIL import Image
import random

root = Tk()
root.title("What Pkmn GOLD starter should you choose?")
root.geometry("500x300")
root.resizable(0,0)
canvas= Canvas(root, width = 250, height = 250)
canvas.pack()
img = PhotoImage(file="squirtle.png", width=250)
img2 = PhotoImage(file="bulbasaur.png")
img3 = PhotoImage(file="charmander.png")

def randoms(event):
    wb = load_workbook("pokedex.xlsx")
    ws = wb.active
    rangeline = ws["C2":"C4"] #creating a list
    pkmn = [] #empty variable that takes list
    for items in rangeline:
        for subitems in items:
            pkmn.append(subitems.value)

    choices = random.choice(pkmn)
    if choices == 'Squirtle':
        root.config(bg='teal')
        canvas.create_image(20, 20, anchor=NW, image=img)
        print("Your starter is " + str(choices))
    elif choices == 'Bulbasaur':
        root.config(bg='lime')
        canvas.create_image(20,20, anchor=NW, image=img2)
        print("Your starter is " + str(choices))
    else:
        root.config(bg='orange')
        canvas.create_image(20,20, anchor=NW, image=img3)
        print("Your starter is " + str(choices))

    updateDisplay(choices)

def updateDisplay(abc):
    displayVariable.set(abc)

button1 = Button(root, text ='Pick!')
button1.bind("<Button-1>", randoms)
button1.pack()
displayVariable = StringVar()
displayLabel =Label(root, textvariable = displayVariable)
displayLabel.pack()

mainloop()